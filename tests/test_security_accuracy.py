import asyncio
import base64
import csv
import hashlib
import hmac
import io
import json
import math
import os
import re
import sqlite3
import socket
import tempfile
import time
import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional
from unittest.mock import AsyncMock, patch

import httpx
from openpyxl import load_workbook
from fastapi import HTTPException
from fastapi.testclient import TestClient

import app.Code.analytics as analytics_module
import app.Code.holdings as holdings_module
import app.Code.utils as utils_module
from casparser.enums import CASFileType, FileType, TransactionType
from casparser.parsers.utils import cas2csv, cas2csv_summary
from casparser.types import (
    CASData,
    Folio,
    InvestorInfo as CasInvestorInfo,
    Scheme,
    SchemeValuation,
    StatementPeriod,
    TransactionData,
)
from app.Code.analytics import (
    _pseudonymize_identifier,
    _sanitize_text,
    get_admin_overview,
    init_analytics_db,
    record_analysis_run,
)
from app.Code.cas_parser import convert_to_excel, parse_with_casparser, recursive_to_dict
from app.Code.env_loader import _should_set_env, load_local_env
from app.Code.billing import (
    AccessStatus,
    CreditReservation,
    _parse_access_status,
    _parse_credit_reservation,
    apply_subscription_status,
    create_razorpay_subscription,
    extract_subscription_event,
    verify_checkout_signature,
    verify_webhook_signature,
)
from app.Code.pdfminer_hardening import (
    ORIGINAL_CMAP_LOADER_ATTR,
    harden_pdfminer_cmap_loading,
)
from app.Code.main import (
    _benchmark_nav_for_date,
    _build_content_security_policy,
    _current_holding_entry_date,
    _get_pdf_parse_executor,
    _get_pdf_parse_timeout_seconds,
    _normalize_amfi_code,
    _parse_amount,
    _parse_iso_date,
    _parse_pdf_upload_direct,
    _parse_pdf_upload_in_subprocess,
    _parse_pdf_upload,
    _prepare_benchmark_history,
    _validate_cas_json_shape,
    _validate_upload,
    app,
    map_casparser_to_analysis,
    MAX_CAS_TRANSACTIONS,
    AnalysisResponse,
    Holding,
    _safe_analysis_response,
)
from app.Code.supabase_auth import _get_supabase_url, _is_admin_user, _is_allowed_supabase_url, _sanitize_username
from app.Code.benchmarks.resolver import resolve_benchmark
from app.Code.utils import calculate_xirr, fetch_live_nav, fetch_nav_history, _parse_amfi_nav_history_text_for_code


class _FakeSupabaseUser:
    user_id = "user_test"
    username = "test-user"
    email = "test@example.com"
    app_metadata = {"role": "admin"}
    user_metadata = {"username": "test-user"}
    is_admin = True


def _test_access_status(
    *,
    can_analyze: bool = True,
    has_unlimited_reports: bool = False,
    cas_report_limit: int = 1,
    cas_reports_used: int = 0,
    remaining_free_reports: int = 1,
    subscription_status: str = "free",
    razorpay_subscription_id: Optional[str] = None,
):
    return AccessStatus(
        can_analyze=can_analyze,
        has_unlimited_reports=has_unlimited_reports,
        cas_report_limit=cas_report_limit,
        cas_reports_used=cas_reports_used,
        remaining_free_reports=remaining_free_reports,
        subscription_status=subscription_status,
        razorpay_subscription_id=razorpay_subscription_id,
        current_period_end=None,
    )


def _test_credit_reservation(*, credit_consumed: bool = True) -> CreditReservation:
    return CreditReservation(
        allowed=True,
        credit_consumed=credit_consumed,
        access=_test_access_status(
            cas_reports_used=1 if credit_consumed else 0,
            remaining_free_reports=0 if credit_consumed else 1,
        ),
    )


async def _fake_require_supabase_user(*args, **kwargs):
    return _FakeSupabaseUser()


class _FakeResponse:
    def __init__(self, status_code: int, payload=None):
        self.status_code = status_code
        self._payload = payload or {}

    def json(self):
        return self._payload


class _FailingAsyncClient:
    def __init__(self, *args, **kwargs):
        pass

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return False

    async def get(self, url):
        raise RuntimeError("timeout")


class _UploadStub:
    def __init__(self, filename: str, content_type: str):
        self.filename = filename
        self.content_type = content_type


def _sleeping_pdf_parse_worker(_content, _password, _result_queue):
    time.sleep(5)


def _successful_pdf_parse_worker(_content, _password, result_sink):
    result_sink.send({"success": True, "data": {"folios": []}})
    result_sink.close()


class TestSecurityAccuracy(unittest.IsolatedAsyncioTestCase):
    async def test_map_returns_new_summary_fields(self):
        fixture = Path("tests/fixtures/sample_cas.json")
        cas_data = json.loads(fixture.read_text(encoding="utf-8"))

        async def fake_live_nav(_):
            return 0.0

        async def fake_benchmark_history(_):
            return {
                "01-01-2023": 100.0,
                "02-01-2023": 100.2,
                "01-01-2024": 120.0,
            }

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_benchmark_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        self.assertIsNotNone(response.summary)
        summary = response.summary
        assert summary is not None
        self.assertEqual(summary.valuation_mode, "live_nav")
        self.assertAlmostEqual(summary.total_market_value, summary.statement_market_value, places=2)
        self.assertAlmostEqual(summary.live_nav_delta_value, 0.0, places=2)
        self.assertIsNotNone(summary.data_coverage)
        self.assertTrue(isinstance(summary.warnings, list))
        self.assertIsNotNone(summary.tax)
        self.assertEqual(summary.tax.equity_ltcg_rate_pct, 12.5)

    def test_analysis_response_sanitizes_non_json_float_values(self):
        response = AnalysisResponse(
            success=True,
            holdings=[
                Holding(
                    fund_family="Test AMC",
                    folio="1/1",
                    scheme_name="Test Fund",
                    amfi="100001",
                    units=math.nan,
                    nav=math.inf,
                    market_value=1000.0,
                    cost_value=900.0,
                    category="Equity",
                    sub_category="Large-Cap",
                    xirr=math.inf,
                    benchmark_xirr=math.nan,
                )
            ],
        )

        safe_response = _safe_analysis_response(response)

        self.assertEqual(safe_response.holdings[0].units, 0.0)
        self.assertEqual(safe_response.holdings[0].nav, 0.0)
        self.assertIsNone(safe_response.holdings[0].xirr)
        self.assertIsNone(safe_response.holdings[0].benchmark_xirr)

    async def test_taxable_gains_apply_ltcg_exemption(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Test Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 200.0, "value": 200000.0, "cost": 100000.0},
                            "transactions": [{"date": "2020-01-01", "amount": 100000.0, "description": "Purchase"}],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2020", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_benchmark_history(_):
            return {"01-01-2020": 100.0, "01-01-2026": 110.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_benchmark_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None
        self.assertEqual(summary.tax.long_term_gains, 100000.0)
        self.assertEqual(summary.tax.short_term_gains, 0.0)
        self.assertEqual(summary.tax.taxable_gains, 0.0)
        self.assertEqual(summary.tax.estimated_tax_liability, 0.0)

    async def test_holding_benchmark_xirr_uses_current_position_window(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Test Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 180.0, "value": 180000.0, "cost": 150000.0},
                            "transactions": [
                                {"date": "2020-01-01", "amount": 100000.0, "units": 1000.0, "description": "Purchase"},
                                {"date": "2022-01-01", "amount": 110000.0, "units": -1000.0, "description": "Redemption"},
                                {"date": "2024-01-01", "amount": 150000.0, "units": 1000.0, "description": "Purchase"},
                            ],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2020", "to": "01-Jan-2026"},
        }

        as_of_str = datetime.now().strftime("%d-%m-%Y")

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(code):
            if str(code) == "100001":
                return {
                    "01-01-2020": 100.0,
                    "01-01-2022": 110.0,
                    "01-01-2024": 150.0,
                    as_of_str: 180.0,
                }
            return {
                "01-01-2020": 100.0,
                "01-01-2022": 110.0,
                "01-01-2024": 130.0,
                as_of_str: 140.0,
            }

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holding = response.holdings[0]
        self.assertEqual(holding.date_of_entry, "2024-01-01")
        self.assertIsNotNone(holding.benchmark_xirr)

        expected_benchmark_value = (150000.0 / 130.0) * 140.0
        expected_benchmark_xirr = calculate_xirr(
            [datetime(2024, 1, 1), datetime.now()],
            [-150000.0, expected_benchmark_value],
        )
        self.assertIsNotNone(expected_benchmark_xirr)
        self.assertAlmostEqual(holding.benchmark_xirr, round(expected_benchmark_xirr or 0.0, 2), places=1)

    async def test_idcw_zero_unit_payout_does_not_zero_out_benchmark_terminal_value(self):
        cas_data = {
            "folios": [
                {
                    "amc": "HDFC Mutual Fund",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "HDFC ELSS Tax saver - Regular Plan - IDCW",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 120.0, "value": 120000.0, "cost": 100000.0},
                            "transactions": [
                                {"date": "2020-01-01", "amount": 100000.0, "units": 1000.0, "description": "Purchase"},
                                {"date": "2022-01-01", "amount": 150000.0, "units": 0.0, "description": "IDCW PAYOUT"},
                            ],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2020", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(code):
            if str(code) == "100001":
                return {"01-01-2020": 100.0, "01-01-2022": 110.0, "01-01-2026": 120.0}
            if str(code) == "152731":
                return {"01-01-2020": 100.0, "01-01-2022": 150.0, "01-01-2026": 120.0}
            return {}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holding = response.holdings[0]
        self.assertEqual(holding.benchmark_name, "Nifty 500 Total Return Index")
        self.assertIsNotNone(holding.benchmark_xirr)

    async def test_summary_cas_warns_and_estimates_performance_from_cost_snapshot(self):
        cas_data = {
            "cas_type": "SUMMARY",
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Test Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 130.0, "value": 130000.0, "cost": 100000.0},
                            "transactions": [],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "2020-01-01", "to": "2026-01-01"},
        }

        async def fake_live_nav(_):
            return 130.0

        async def fake_nav_history(code):
            if str(code) == "100001":
                return {"01-01-2020": 100.0, "01-01-2026": 130.0}
            return {"01-01-2020": 100.0, "01-01-2026": 125.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None
        self.assertEqual(summary.analysis_version, "2026.06")
        warning_codes = {warning.code for warning in summary.warnings}
        self.assertIn("CAS_NO_TRANSACTIONS", warning_codes)
        self.assertIn("PERFORMANCE_ESTIMATED_SNAPSHOT", warning_codes)
        holding = response.holdings[0]
        self.assertEqual(holding.performance_source, "estimated_snapshot")
        self.assertIsNotNone(holding.xirr)
        self.assertIsNotNone(holding.date_of_entry)

    async def test_performance_uses_distinct_1y_and_3y_windows(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Test Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 130.0, "value": 130000.0, "cost": 100000.0},
                            "transactions": [{"date": "2023-01-01", "amount": 100000.0, "units": 1000.0, "description": "Purchase"}],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2023", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(code):
            if str(code) == "100001":
                return {
                    "01-01-2023": 100.0,
                    "25-02-2025": 120.0,
                    "25-02-2026": 130.0,
                }
            # Nifty 50 proxy used by equity benchmark mapping.
            return {
                "01-01-2023": 100.0,
                "25-02-2025": 110.0,
                "25-02-2026": 125.0,
            }

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None and summary.performance_summary is not None
        one_year = summary.performance_summary.one_year.underperforming_pct
        three_year = summary.performance_summary.three_year.underperforming_pct
        self.assertNotEqual(one_year, three_year)
        # Current fund value exceeds benchmark value for same cashflow path.
        self.assertLess((response.holdings[0].missed_gains or 0.0), 0.0)

    async def test_tax_lot_losses_are_set_off_before_tax(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Winning Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 400.0, "value": 400000.0, "cost": 100000.0},
                            "transactions": [{"date": "2020-01-01", "amount": 100000.0, "units": 1000.0, "description": "Purchase"}],
                        },
                        {
                            "scheme": "Losing Equity Fund",
                            "amfi": "100002",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 100.0, "value": 100000.0, "cost": 300000.0},
                            "transactions": [{"date": "2020-01-01", "amount": 300000.0, "units": 1000.0, "description": "Purchase"}],
                        },
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2020", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {"01-01-2020": 100.0, "25-02-2026": 120.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None
        self.assertEqual(summary.tax.long_term_gains, 100000.0)
        self.assertEqual(summary.tax.taxable_gains, 0.0)
        self.assertEqual(summary.tax.estimated_tax_liability, 0.0)

    async def test_overlap_absent_when_real_holdings_unavailable(self):
        fixture = Path("tests/fixtures/sample_cas.json")
        cas_data = json.loads(fixture.read_text(encoding="utf-8"))

        async def fake_live_nav(_):
            return 0.0

        async def fake_benchmark_history(_):
            return {"01-01-2023": 100.0, "01-01-2024": 120.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_benchmark_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertIsNone(response.summary.overlap)
        self.assertEqual(response.summary.data_coverage.overlap_source, "none")

    def test_benchmark_fallback_uses_nearest_previous_date(self):
        history, keys, days, _ = _prepare_benchmark_history(
            {"01-01-2023": 100.0, "03-01-2023": 103.0}
        )
        nav, is_exact = _benchmark_nav_for_date("2023-01-02", history, keys, days)
        self.assertEqual(nav, 100.0)
        self.assertFalse(is_exact)

    def test_current_holding_entry_date_uses_remaining_units_after_full_exit(self):
        entry_dt = _current_holding_entry_date(
            50.0,
            [
                (datetime(2020, 1, 1), 100.0, 10000.0),
                (datetime(2021, 1, 1), -100.0, 12000.0),
                (datetime(2024, 1, 1), 50.0, 7000.0),
            ],
            datetime(2020, 1, 1),
        )
        self.assertEqual(entry_dt, datetime(2024, 1, 1))

    def test_benchmark_mapping_uses_sebi_tier1_proxies(self):
        flexi = resolve_benchmark(None, "Test Flexi Cap Fund - Direct Plan - Growth", scheme_type="EQUITY")
        self.assertEqual(flexi.benchmark_name, "Nifty 500 Total Return Index")
        self.assertEqual([c.code for c in flexi.components], ["152731"])

        nasdaq = resolve_benchmark(
            None,
            "ICICI Prudential NASDAQ 100 Index Fund - Direct Plan - Growth",
            scheme_type="EQUITY",
        )
        self.assertEqual(nasdaq.benchmark_name, "Nasdaq 100 Total Return Index")
        self.assertEqual([c.code for c in nasdaq.components], ["149219"])

        gold = resolve_benchmark(None, "HDFC Gold ETF Fund of Fund - Direct Plan", scheme_type="ETF")
        self.assertEqual(gold.benchmark_name, "Domestic Price of Gold")
        self.assertEqual([c.code for c in gold.components], ["119132"])

    def test_sectoral_equity_funds_use_sector_tier1_benchmarks(self):
        sectoral = resolve_benchmark(
            None,
            "Test Banking and Financial Services Fund",
            scheme_type="EQUITY",
        )
        self.assertEqual(sectoral.benchmark_name, "Nifty Bank Total Return Index")
        self.assertEqual([c.code for c in sectoral.components], ["147620"])

    def test_technology_and_infrastructure_funds_use_sector_benchmarks(self):
        technology = resolve_benchmark(None, "ICICI Prudential Technology Fund - Growth", scheme_type="EQUITY")
        self.assertEqual(technology.benchmark_name, "Nifty IT Total Return Index")
        self.assertEqual([c.code for c in technology.components], ["153322"])

        infrastructure = resolve_benchmark(None, "UTI Infrastructure Fund - Regular Plan", scheme_type="EQUITY")
        self.assertEqual(infrastructure.benchmark_name, "Nifty Infrastructure Total Return Index")
        self.assertEqual([c.code for c in infrastructure.components], ["153078"])

    def test_business_cycle_equity_funds_use_bse_500_benchmark(self):
        business_cycle = resolve_benchmark(
            None,
            "Motilal Oswal Business Cycle Fund - Direct Plan - Growth",
            scheme_type="EQUITY",
        )
        self.assertEqual(business_cycle.benchmark_name, "S&P BSE 500 Total Return Index")
        self.assertEqual([c.code for c in business_cycle.components], ["151728"])

    async def test_analysis_uses_updated_benchmark_name_for_flexi_cap_funds(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Parag Parikh Flexi Cap Fund",
                            "amfi": "122639",
                            "type": "EQUITY",
                            "close": 100.0,
                            "valuation": {"nav": 120.0, "value": 12000.0, "cost": 10000.0},
                            "transactions": [{"date": "2024-01-01", "amount": 10000.0, "description": "Purchase"}],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2025"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {"01-01-2024": 100.0, "01-01-2025": 110.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        self.assertEqual(response.holdings[0].benchmark_name, "Nifty 500 Total Return Index")
        self.assertEqual(response.holdings[0].sebi_category, "flexi_cap")
        self.assertEqual(response.holdings[0].benchmark_source, "sebi_tier1")

    async def test_analysis_labels_technology_and_infrastructure_benchmarks_without_proxy_history(self):
        cas_data = {
            "folios": [
                {
                    "amc": "ICICI Prudential Mutual Fund",
                    "folio": "20644404/27",
                    "schemes": [
                        {
                            "scheme": "ICICI Prudential Technology Fund - Growth",
                            "amfi": "100363",
                            "type": "EQUITY",
                            "close": 100.0,
                            "valuation": {"nav": 120.0, "value": 12000.0, "cost": 10000.0},
                            "transactions": [{"date": "2024-01-01", "amount": 10000.0, "description": "Purchase"}],
                        }
                    ],
                },
                {
                    "amc": "UTI Mutual Fund",
                    "folio": "511213501963/0",
                    "schemes": [
                        {
                            "scheme": "UTI - Infrastructure Fund",
                            "amfi": "102395",
                            "type": "EQUITY",
                            "close": 100.0,
                            "valuation": {"nav": 130.0, "value": 13000.0, "cost": 10000.0},
                            "transactions": [{"date": "2024-01-01", "amount": 10000.0, "description": "Purchase"}],
                        }
                    ],
                },
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2025"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(amfi_code):
            if amfi_code in {"100001", "100002"}:
                return {"01-01-2024": 100.0, "01-01-2025": 110.0}
            return {}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holdings_by_name = {holding.scheme_name: holding for holding in response.holdings}
        self.assertEqual(
            holdings_by_name["ICICI Prudential Technology Fund - Growth"].benchmark_name,
            "Nifty IT Total Return Index",
        )
        self.assertIn("Sectoral", holdings_by_name["ICICI Prudential Technology Fund - Growth"].sub_category)
        self.assertEqual(
            holdings_by_name["UTI - Infrastructure Fund"].benchmark_name,
            "Nifty Infrastructure Total Return Index",
        )
        self.assertIn("Sectoral", holdings_by_name["UTI - Infrastructure Fund"].sub_category)

    async def test_performance_summary_exposes_comparable_coverage(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Parag Parikh Flexi Cap Fund",
                            "amfi": "122639",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 100.0, "value": 100000.0, "cost": 90000.0},
                            "transactions": [{"date": "2024-01-01", "amount": 90000.0, "units": 1000.0, "description": "Purchase"}],
                        },
                        {
                            "scheme": "Test Banking and Financial Services Fund",
                            "amfi": "100002",
                            "type": "EQUITY",
                            "close": 1000.0,
                            "valuation": {"nav": 100.0, "value": 100000.0, "cost": 90000.0},
                            "transactions": [{"date": "2024-01-01", "amount": 90000.0, "units": 1000.0, "description": "Purchase"}],
                        },
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(code):
            if str(code) in {"100001", "100002"}:
                return {
                    "01-01-2024": 90.0,
                    "01-01-2025": 95.0,
                    "01-01-2026": 100.0,
                }
            return {
                "01-01-2024": 100.0,
                "01-01-2025": 105.0,
                "01-01-2026": 110.0,
            }

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None and summary.performance_summary is not None
        self.assertEqual(summary.performance_summary.one_year.comparable_pct, 100.0)

    async def test_analysis_uses_remaining_holding_entry_date_and_parses_statement_cost(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Parag Parikh Flexi Cap Fund",
                            "amfi": "122639",
                            "type": "EQUITY",
                            "close": 50.0,
                            "valuation": {"nav": 140.0, "value": 7000.0, "cost": "7,000.0"},
                            "transactions": [
                                {"date": "2020-01-01", "amount": 10000.0, "units": 100.0, "description": "Purchase"},
                                {"date": "2021-01-01", "amount": 12000.0, "units": -100.0, "description": "Redemption"},
                                {"date": "2024-01-01", "amount": 7000.0, "units": 50.0, "description": "Purchase"},
                            ],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2020", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {
                "01-01-2020": 100.0,
                "01-01-2021": 110.0,
                "01-01-2024": 120.0,
                "01-01-2026": 130.0,
            }

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holding = response.holdings[0]
        self.assertEqual(holding.date_of_entry, "2024-01-01")
        self.assertEqual(holding.cost_value, 7000.0)

    async def test_partial_redemption_reconstructs_remaining_cost_basis(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Partial Redemption Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 50.0,
                            "valuation": {"nav": 100.0, "value": 5000.0},
                            "transactions": [
                                {"date": "2024-01-01", "amount": 10000.0, "units": 100.0, "description": "Purchase"},
                                {"date": "2025-01-01", "amount": 7500.0, "units": -50.0, "description": "Redemption"},
                            ],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {"01-01-2024": 100.0, "01-01-2025": 110.0, "01-01-2026": 100.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holding = response.holdings[0]
        self.assertEqual(holding.cost_value, 5000.0)
        self.assertEqual(holding.gain_loss, 0.0)
        self.assertEqual(response.summary.total_cost_value, 5000.0)

    async def test_analysis_parses_comma_formatted_numeric_strings(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Formatted Numbers Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": "70",
                            "valuation": {"nav": "100.50", "value": "7,035.00", "cost": "6,000.00"},
                            "transactions": [{"date": "2024-01-01", "amount": "6,000.00", "units": "70", "description": "Purchase"}],
                        }
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2026"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {"01-01-2024": 100.0, "01-01-2026": 100.5}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        holding = response.holdings[0]
        self.assertEqual(holding.units, 70.0)
        self.assertEqual(holding.nav, 100.5)
        self.assertEqual(holding.market_value, 7035.0)
        self.assertEqual(response.summary.statement_market_value, 7035.0)

    def test_eval_removed_from_holdings_cache_loader(self):
        holdings_source = Path("app/Code/holdings.py").read_text(encoding="utf-8")
        self.assertNotIn("eval(", holdings_source)

    async def test_debt_type_not_misclassified_by_growth_keyword(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Test Equity Fund - Growth",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 10.0,
                            "valuation": {"nav": 100.0, "value": 1000.0, "cost": 800.0},
                            "transactions": [{"date": "2024-01-01", "amount": 800.0, "description": "Purchase"}],
                        },
                        {
                            "scheme": "Test Short Duration Fund - Growth",
                            "amfi": "100002",
                            "type": "DEBT",
                            "close": 20.0,
                            "valuation": {"nav": 100.0, "value": 2000.0, "cost": 1800.0},
                            "transactions": [{"date": "2024-01-01", "amount": 1800.0, "description": "Purchase"}],
                        },
                    ],
                }
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2025"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_benchmark_history(_):
            return {"01-01-2024": 100.0, "01-01-2025": 110.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_benchmark_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None
        self.assertGreater(summary.equity_pct, 0.0)
        self.assertLess(summary.equity_pct, 100.0)
        self.assertIsNotNone(summary.fixed_income)
        assert summary.fixed_income is not None
        self.assertGreater(summary.fixed_income.current_value, 0.0)

    async def test_concentration_top_funds_aggregate_same_scheme_across_folios(self):
        cas_data = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Repeated Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 10.0,
                            "valuation": {"nav": 100.0, "value": 1000.0, "cost": 900.0},
                            "transactions": [{"date": "2024-01-01", "amount": 900.0, "units": 10.0, "description": "Purchase"}],
                        }
                    ],
                },
                {
                    "amc": "Test AMC",
                    "folio": "2/2",
                    "schemes": [
                        {
                            "scheme": "Repeated Equity Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 20.0,
                            "valuation": {"nav": 100.0, "value": 2000.0, "cost": 1800.0},
                            "transactions": [{"date": "2024-01-01", "amount": 1800.0, "units": 20.0, "description": "Purchase"}],
                        },
                        {
                            "scheme": "Smaller Equity Fund",
                            "amfi": "100002",
                            "type": "EQUITY",
                            "close": 15.0,
                            "valuation": {"nav": 100.0, "value": 1500.0, "cost": 1500.0},
                            "transactions": [{"date": "2024-01-01", "amount": 1500.0, "units": 15.0, "description": "Purchase"}],
                        },
                    ],
                },
            ],
            "statement_period": {"from": "01-Jan-2024", "to": "01-Jan-2025"},
        }

        async def fake_live_nav(_):
            return 0.0

        async def fake_nav_history(_):
            return {"01-01-2024": 100.0, "01-01-2025": 100.0}

        with patch("app.Code.main.fetch_live_nav", new=fake_live_nav), patch(
            "app.Code.main.fetch_nav_history", new=fake_nav_history
        ), patch("app.Code.main.save_cache_async", new=AsyncMock()), patch(
            "app.Code.main.get_holdings_for_schemes", new=AsyncMock(return_value={})
        ), patch(
            "app.Code.main.save_amfi_cache_async", new=AsyncMock()
        ):
            response = await map_casparser_to_analysis(cas_data)

        self.assertTrue(response.success)
        summary = response.summary
        assert summary is not None
        self.assertEqual(summary.concentration.fund_count, 2)
        self.assertEqual(summary.concentration.top_funds[0].name, "Repeated Equity Fund")
        self.assertEqual(summary.concentration.top_funds[0].value, 3000.0)
        self.assertAlmostEqual(summary.concentration.top_funds[0].allocation_pct, 66.7)


class TestErrorSanitization(unittest.TestCase):
    def test_load_local_env_fills_empty_placeholder_values_from_file(self):
        temp_key = "TEST_ENV_LOADER_EMPTY_KEY"
        original_temp = os.environ.get(temp_key)

        try:
            os.environ[temp_key] = ""
            with tempfile.TemporaryDirectory() as tmpdir:
                env_path = Path(tmpdir) / ".env"
                env_path.write_text(f"{temp_key}=loaded-value\n", encoding="utf-8")
                load_local_env(env_path)

            self.assertEqual(os.environ.get(temp_key), "loaded-value")
        finally:
            if original_temp is None:
                os.environ.pop(temp_key, None)
            else:
                os.environ[temp_key] = original_temp

    def test_load_local_env_sets_missing_values_without_overwriting_existing_ones(self):
        temp_key = "TEST_ENV_LOADER_KEY"
        preserved_key = "TEST_ENV_LOADER_PRESERVE"
        original_temp = os.environ.pop(temp_key, None)
        original_preserved = os.environ.get(preserved_key)
        os.environ[preserved_key] = "already-set"

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                env_path = Path(tmpdir) / ".env"
                env_path.write_text(
                    f"{temp_key}=loaded-value\n{preserved_key}=should-not-overwrite\n",
                    encoding="utf-8",
                )

                load_local_env(env_path)

            self.assertEqual(os.environ.get(temp_key), "loaded-value")
            self.assertEqual(os.environ.get(preserved_key), "already-set")
        finally:
            if original_temp is None:
                os.environ.pop(temp_key, None)
            else:
                os.environ[temp_key] = original_temp

            if original_preserved is None:
                os.environ.pop(preserved_key, None)
            else:
                os.environ[preserved_key] = original_preserved

    def test_public_frontend_env_override_is_limited_to_known_public_prefixes(self):
        with patch.dict(
            os.environ,
            {
                "APP_SUPABASE_URL": "https://old.supabase.co",
                "APP_SECRET_VALUE": "server-secret",
            },
            clear=False,
        ):
            self.assertTrue(
                _should_set_env(
                    "APP_SUPABASE_URL",
                    override=False,
                    override_public_frontend=True,
                )
            )
            self.assertFalse(
                _should_set_env(
                    "APP_SECRET_VALUE",
                    override=False,
                    override_public_frontend=True,
                )
            )
        with patch.dict(os.environ, {}, clear=True):
            self.assertFalse(
                _should_set_env(
                    "APP_SECRET_VALUE",
                    override=False,
                    override_public_frontend=True,
                )
            )
            self.assertFalse(
                _should_set_env(
                    "SUPABASE_SERVICE_ROLE_KEY",
                    override=True,
                    override_public_frontend=True,
                )
            )
            self.assertTrue(
                _should_set_env(
                    "APP_SUPABASE_ANON_KEY",
                    override=False,
                    override_public_frontend=True,
                )
            )

    def test_analyze_returns_sanitized_internal_error(self):
        client = TestClient(app)
        files = {"file": ("sample.json", b'{"folios": []}', "application/json")}
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.reserve_analysis_credit",
            new=AsyncMock(return_value=type("Reservation", (), {"credit_consumed": True})()),
        ), patch(
            "app.Code.main.refund_analysis_credit",
            new=AsyncMock(),
        ), patch(
            "app.Code.main.map_casparser_to_analysis",
            side_effect=RuntimeError("boom secret details"),
        ):
            response = client.post("/api/analyze", files=files, data={"password": ""})
        body = response.json()
        self.assertIn("Request ID", body.get("error", ""))
        self.assertNotIn("secret details", body.get("error", ""))

    def test_upload_signature_validation_rejects_invalid_pdf_bytes(self):
        client = TestClient(app)
        files = {"file": ("statement.pdf", b"not-a-real-pdf", "application/pdf")}
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user):
            response = client.post("/api/analyze", files=files, data={"password": ""})
        body = response.json()
        self.assertFalse(body.get("success", True))
        self.assertIn("invalid or corrupted", body.get("error", "").lower())

    def test_upload_validation_accepts_content_type_parameters(self):
        json_error = _validate_upload(
            _UploadStub("sample.json", "application/json; charset=utf-8"),
            b'{"folios": []}',
        )
        pdf_error = _validate_upload(
            _UploadStub("statement.pdf", "application/pdf; charset=binary"),
            b"%PDF-1.7\n",
        )

        self.assertIsNone(json_error)
        self.assertIsNone(pdf_error)

    def test_analyze_rejects_invalid_cas_json_shape(self):
        client = TestClient(app)
        files = {"file": ("sample.json", b'{"folios":["oops"]}', "application/json")}
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user):
            response = client.post("/api/analyze", files=files, data={"password": ""})
        body = response.json()
        self.assertFalse(body.get("success", True))
        self.assertIn("invalid cas json format", body.get("error", "").lower())
        self.assertNotIn("request id", body.get("error", "").lower())

    def test_analyze_rejects_non_finite_numeric_values(self):
        client = TestClient(app)
        payload = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Bad Number Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": "NaN",
                            "valuation": {"nav": 10, "value": 100},
                            "transactions": [],
                        }
                    ],
                }
            ]
        }
        files = {"file": ("sample.json", json.dumps(payload).encode("utf-8"), "application/json")}
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user):
            response = client.post("/api/analyze", files=files, data={"password": ""})

        body = response.json()
        self.assertFalse(body.get("success", True))
        self.assertIn("finite number", body.get("error", "").lower())
        self.assertNotIn("request id", body.get("error", "").lower())

    def test_parse_amount_accepts_common_inr_currency_formats(self):
        self.assertEqual(_parse_amount("Rs. 1,23,456.78"), 123456.78)
        self.assertEqual(_parse_amount("INR 2,500"), 2500.0)
        self.assertEqual(_parse_amount("\u20b93,000.50"), 3000.50)
        self.assertEqual(_parse_amount("(Rs. 1,000.25)"), -1000.25)

    def test_cas_shape_limits_reject_transaction_floods(self):
        payload = {
            "folios": [
                {
                    "amc": "Test AMC",
                    "folio": "1/1",
                    "schemes": [
                        {
                            "scheme": "Large Fund",
                            "amfi": "100001",
                            "type": "EQUITY",
                            "close": 1,
                            "valuation": {"nav": 1, "value": 1},
                            "transactions": [
                                {"date": "2024-01-01", "amount": 1, "units": 1}
                                for _ in range(MAX_CAS_TRANSACTIONS + 1)
                            ],
                        }
                    ],
                }
            ]
        }

        error = _validate_cas_json_shape(payload)

        self.assertIsNotNone(error)
        self.assertIn("too many transactions", error.lower())

    def test_non_numeric_amfi_codes_are_ignored_before_external_fetch(self):
        self.assertEqual(_normalize_amfi_code("100001"), "100001")
        self.assertEqual(_normalize_amfi_code(100001.0), "100001")
        self.assertEqual(_normalize_amfi_code("100001/../../evil"), "")

    def test_nav_fetchers_reject_invalid_amfi_codes_without_network(self):
        with patch("app.Code.utils._get_client", side_effect=AssertionError("network called")):
            self.assertEqual(asyncio.run(fetch_live_nav("100001/../../evil")), 0.0)
            self.assertEqual(asyncio.run(fetch_nav_history("100001/../../evil")), {})

    def test_amfi_history_parser_extracts_target_scheme_rows(self):
        text = (
            "Scheme Code;Scheme Name;ISIN Div Payout/ISIN Growth;ISIN Div Reinvestment;"
            "Net Asset Value;Repurchase Price;Sale Price;Date\n"
            "Open Ended Schemes ( Equity Scheme )\n"
            "Axis Mutual Fund\n"
            "152731;Axis Nifty 500 Index Fund - Direct Plan - Growth Option;INF846K019W9;-;"
            "10.2500;;;01-Jan-2024\n"
            "120716;UTI Nifty 50 Index Fund - Growth Option- Direct;INF789F01XA0;-;"
            "164.8682;;;29-May-2026\n"
            "not-a-code;Bad Row;;;;;;01-Jan-2024\n"
        )

        history = _parse_amfi_nav_history_text_for_code(text, "152731")

        self.assertEqual(history, {"01-01-2024": 10.25})

    def test_nav_history_falls_back_to_official_amfi_history_export(self):
        class TextResponse:
            def __init__(self, status_code, text="", payload=None):
                self.status_code = status_code
                self.text = text
                self._payload = payload or {}

            def json(self):
                return self._payload

        class FakeClient:
            def __init__(self):
                self.calls = []

            async def get(self, url, **kwargs):
                self.calls.append((url, kwargs))
                if "api.mfapi.in" in url:
                    return TextResponse(502, payload={})
                if url == utils_module.NAV_ALL_URL:
                    return TextResponse(
                        200,
                        "Open Ended Schemes(Equity Scheme)\n"
                        "Axis Mutual Fund\n"
                        "152731;INF846K019W9;-;Axis Nifty 500 Index Fund - Direct Plan - Growth Option;"
                        "11.0000;29-May-2026\n",
                    )
                if url == utils_module.AMFI_FUND_LIST_URL:
                    return TextResponse(200, '{"mfId":"53","mfName":"Axis Mutual Fund"}')
                if url == utils_module.NAV_HISTORY_URL:
                    return TextResponse(
                        200,
                        "Scheme Code;Scheme Name;ISIN Div Payout/ISIN Growth;ISIN Div Reinvestment;"
                        "Net Asset Value;Repurchase Price;Sale Price;Date\n"
                        "152731;Axis Nifty 500 Index Fund - Direct Plan - Growth Option;INF846K019W9;-;"
                        "10.0000;;;01-Jan-2024\n",
                    )
                return TextResponse(404)

        async def fake_get_client():
            return fake_client

        fake_client = FakeClient()
        utils_module._history_cache.clear()
        utils_module._history_cache_years.clear()
        utils_module._history_full_cache.clear()
        utils_module._history_primary_failed_codes.clear()
        utils_module._fetch_locks.clear()
        utils_module._navall_map.clear()
        utils_module._navall_scheme_amc_map.clear()
        utils_module._navall_history_date_map.clear()
        utils_module._navall_cache_date = None
        utils_module._amfi_fund_ids.clear()
        utils_module._amfi_history_chunk_cache.clear()
        utils_module._amfi_history_chunk_locks.clear()

        with patch("app.Code.utils._get_client", new=fake_get_client):
            history = asyncio.run(fetch_nav_history("152731", required_dates=[date(2024, 1, 1)]))

        self.assertEqual(history["01-01-2024"], 10.0)
        self.assertEqual(history["29-05-2026"], 11.0)
        history_calls = [kwargs for url, kwargs in fake_client.calls if url == utils_module.NAV_HISTORY_URL]
        self.assertTrue(history_calls)
        self.assertEqual(history_calls[0]["params"].get("mf"), "53")

    def test_parse_amount_rejects_non_finite_values(self):
        self.assertIsNone(_parse_amount("NaN"))
        self.assertIsNone(_parse_amount("Infinity"))
        self.assertEqual(_parse_amount("Rs 1,234.50"), 1234.5)

    def test_parse_iso_date_accepts_iso_datetime_values(self):
        self.assertEqual(_parse_iso_date("2024-01-02T10:30:00").date(), date(2024, 1, 2))
        self.assertEqual(_parse_iso_date("2024-01-02T10:30:00Z").date(), date(2024, 1, 2))
        self.assertEqual(_parse_iso_date(date(2024, 1, 2)).date(), date(2024, 1, 2))

    def test_analyze_rejects_oversized_upload_before_processing(self):
        client = TestClient(app)
        with patch("app.Code.main.MAX_UPLOAD_BYTES", 10):
            files = {"file": ("sample.json", b'{"folios": []}', "application/json")}
            with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user):
                response = client.post("/api/analyze", files=files, data={"password": ""})
        body = response.json()
        self.assertFalse(body.get("success", True))
        self.assertIn("too large", body.get("error", "").lower())

    def test_parse_pdf_json_serializes_parser_decimal_and_date_values(self):
        client = TestClient(app)
        parser_payload = {
            "folios": [
                {
                    "amount": Decimal("100.50"),
                    "date": date(2024, 1, 2),
                    "type": TransactionType.PURCHASE,
                }
            ]
        }

        reserve_mock = AsyncMock(return_value=_test_credit_reservation())
        refund_mock = AsyncMock()
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.reserve_analysis_credit",
            new=reserve_mock,
        ), patch(
            "app.Code.main.refund_analysis_credit",
            new=refund_mock,
        ), patch(
            "app.Code.main._parse_pdf_upload",
            new=AsyncMock(return_value={"success": True, "data": parser_payload}),
        ):
            response = client.post(
                "/api/parse_pdf",
                files={"file": ("statement.pdf", b"%PDF-1.7\n", "application/pdf")},
                data={"password": "", "output_format": "json"},
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["folios"][0]["amount"], 100.5)
        self.assertEqual(body["folios"][0]["date"], "2024-01-02")
        self.assertEqual(body["folios"][0]["type"], TransactionType.PURCHASE.value)
        reserve_mock.assert_awaited_once_with("user_test", is_admin=True)
        refund_mock.assert_not_awaited()

    def test_parse_pdf_rejects_malformed_parser_payload_before_export(self):
        client = TestClient(app)
        malformed_payload = {"folios": "not-a-list"}

        reserve_mock = AsyncMock(return_value=_test_credit_reservation())
        refund_mock = AsyncMock()
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.reserve_analysis_credit",
            new=reserve_mock,
        ), patch(
            "app.Code.main.refund_analysis_credit",
            new=refund_mock,
        ), patch(
            "app.Code.main._parse_pdf_upload",
            new=AsyncMock(return_value={"success": True, "data": malformed_payload}),
        ):
            response = client.post(
                "/api/parse_pdf",
                files={"file": ("statement.pdf", b"%PDF-1.7\n", "application/pdf")},
                data={"password": "", "output_format": "excel"},
            )

        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertIn("invalid cas json format", body.get("error", "").lower())
        self.assertNotIn("request id", body.get("error", "").lower())
        reserve_mock.assert_awaited_once_with("user_test", is_admin=True)
        refund_mock.assert_awaited_once_with("user_test")

    def test_parse_pdf_requires_report_credit_before_processing(self):
        client = TestClient(app)
        parse_mock = AsyncMock(return_value={"success": True, "data": {"folios": []}})

        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.reserve_analysis_credit",
            new=AsyncMock(side_effect=HTTPException(status_code=402, detail="Report limit reached.")),
        ), patch(
            "app.Code.main._parse_pdf_upload",
            new=parse_mock,
        ):
            response = client.post(
                "/api/parse_pdf",
                files={"file": ("statement.pdf", b"%PDF-1.7\n", "application/pdf")},
                data={"password": "", "output_format": "json"},
            )

        self.assertEqual(response.status_code, 402)
        self.assertEqual(response.json().get("detail"), "Report limit reached.")
        parse_mock.assert_not_awaited()

    def test_parse_pdf_refunds_report_credit_on_parse_failure(self):
        client = TestClient(app)
        refund_mock = AsyncMock()

        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.reserve_analysis_credit",
            new=AsyncMock(return_value=_test_credit_reservation()),
        ), patch(
            "app.Code.main.refund_analysis_credit",
            new=refund_mock,
        ), patch(
            "app.Code.main._parse_pdf_upload",
            new=AsyncMock(return_value={"success": False, "error": "Incorrect password."}),
        ):
            response = client.post(
                "/api/parse_pdf",
                files={"file": ("statement.pdf", b"%PDF-1.7\n", "application/pdf")},
                data={"password": "", "output_format": "json"},
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("incorrect password", response.json().get("error", "").lower())
        refund_mock.assert_awaited_once_with("user_test")

    def test_spa_routes_return_index_html_locally_with_admin_api_protected(self):
        client = TestClient(app)

        admin_response = client.get("/admin")
        dashboard_response = client.get("/dashboard")
        pricing_response = client.get("/pricing")

        self.assertEqual(admin_response.status_code, 200)
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(pricing_response.status_code, 200)
        self.assertIn("<!doctype html>", admin_response.text.lower())
        self.assertIn("<!doctype html>", dashboard_response.text.lower())
        self.assertIn("<!doctype html>", pricing_response.text.lower())
        self.assertEqual(admin_response.headers.get("cache-control"), "no-store, max-age=0")
        self.assertEqual(dashboard_response.headers.get("cache-control"), "no-store, max-age=0")
        self.assertEqual(pricing_response.headers.get("cache-control"), "no-store, max-age=0")

        admin_api_response = client.get("/api/admin/overview")
        self.assertEqual(admin_api_response.status_code, 401)
        self.assertEqual(admin_api_response.headers.get("cache-control"), "no-store, max-age=0")

    def test_analyze_disables_client_and_proxy_caching(self):
        client = TestClient(app)

        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user):
            response = client.post(
                "/api/analyze",
                files={"file": ("sample.json", b"{}", "application/json")},
                data={"password": ""},
            )

        self.assertEqual(response.headers.get("cache-control"), "no-store, max-age=0")
        self.assertEqual(response.headers.get("pragma"), "no-cache")
        self.assertEqual(response.headers.get("expires"), "0")
        self.assertEqual(response.headers.get("x-content-type-options"), "nosniff")
        self.assertEqual(response.headers.get("x-frame-options"), "DENY")
        csp = response.headers.get("content-security-policy", "")
        self.assertIn("default-src 'self'", csp)
        self.assertIn("https://*.supabase.co", csp)
        self.assertNotIn("https://*.supabase.com", csp)
        self.assertIn("object-src 'none'", csp)
        self.assertIn("frame-ancestors 'none'", csp)

    def test_csp_rejects_non_supabase_remote_connect_sources(self):
        with patch.dict(
            os.environ,
            {
                "SUPABASE_URL": "https://example.com",
                "APP_SUPABASE_URL": "https://project.supabase.co",
            },
            clear=False,
        ):
            csp = _build_content_security_policy()

        self.assertNotIn("https://example.com", csp)
        self.assertIn("https://project.supabase.co", csp)

    def test_csp_preserves_allowed_supabase_ports_and_razorpay_checkout_sources(self):
        with patch.dict(
            os.environ,
            {
                "SUPABASE_URL": "http://127.0.0.1:54321",
                "APP_SUPABASE_URL": "http://[::1]:54322",
            },
            clear=False,
        ):
            csp = _build_content_security_policy()

        self.assertIn("http://127.0.0.1:54321", csp)
        self.assertIn("http://[::1]:54322", csp)
        self.assertIn("https://api.razorpay.com", csp)
        self.assertIn("https://checkout.razorpay.com", csp)
        self.assertIn("https://lumberjack.razorpay.com", csp)
        self.assertIn("https://lumberjack-metrics.razorpay.com", csp)
        self.assertIn("https://checkout-static-next.razorpay.com", csp)
        self.assertIn("https://cdn.razorpay.com", csp)

    def test_vercel_spa_routes_are_served_by_fastapi_for_security_headers(self):
        vercel_config = json.loads(Path("vercel.json").read_text(encoding="utf-8"))
        routes = {route["src"]: route for route in vercel_config["routes"]}

        for route in (
            "/",
            "/dashboard",
            "/dashboard/(.*)",
            "/admin",
            "/admin/(.*)",
            "/pricing",
            "/pricing/(.*)",
        ):
            self.assertEqual(routes[route]["dest"], "app/Code/main.py")

    def test_public_test_endpoint_is_not_exposed(self):
        client = TestClient(app)

        response = client.get("/test")

        self.assertEqual(response.status_code, 404)

class TestAnalyticsPrivacyHelpers(unittest.TestCase):
    def test_pseudonymize_identifier_hashes_stably_without_raw_id(self):
        first = _pseudonymize_identifier("user_123", "usr")
        second = _pseudonymize_identifier("user_123", "usr")

        self.assertEqual(first, second)
        self.assertRegex(first or "", r"^usr_[0-9a-f]{16}$")
        self.assertNotIn("user_123", first or "")

    def test_sanitize_text_redacts_phone_and_normalizes_whitespace(self):
        sanitized = _sanitize_text(" Call me at +91 98765 43210 \nthanks ")

        self.assertEqual(sanitized, "Call me at [REDACTED_PHONE] thanks")

    def test_analysis_telemetry_does_not_persist_or_return_portfolio_value(self):
        original_db_path = analytics_module.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                analytics_module.DB_PATH = Path(tmp_dir) / "analytics.db"
                init_analytics_db()

                record_analysis_run(
                    request_id="req_privacy",
                    user_id="user_123",
                    username="privacy-test",
                    session_id=None,
                    file_type="json",
                    status="success",
                    duration_ms=123,
                    holdings_count=4,
                    total_market_value=9876543.21,
                )

                overview = get_admin_overview()
                self.assertEqual(len(overview["recent_analyses"]), 1)
                self.assertNotIn("total_market_value", overview["recent_analyses"][0])
                self.assertNotIn("total_market_value", overview["recent_logs"][0]["metadata"])

                conn = sqlite3.connect(analytics_module.DB_PATH)
                try:
                    stored_value = conn.execute(
                        "SELECT total_market_value FROM analysis_runs WHERE request_id = ?",
                        ("req_privacy",),
                    ).fetchone()[0]
                finally:
                    conn.close()
                self.assertIsNone(stored_value)
        finally:
            analytics_module.DB_PATH = original_db_path

    def test_init_analytics_db_migrates_legacy_optional_columns(self):
        original_db_path = analytics_module.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                analytics_module.DB_PATH = Path(tmp_dir) / "legacy_analytics.db"
                conn = sqlite3.connect(analytics_module.DB_PATH)
                try:
                    conn.execute(
                        """
                        CREATE TABLE analysis_runs (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            request_id TEXT NOT NULL,
                            user_id TEXT NOT NULL,
                            status TEXT NOT NULL,
                            created_at TEXT NOT NULL
                        )
                        """
                    )
                    conn.execute(
                        """
                        CREATE TABLE audit_logs (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            user_id TEXT,
                            route TEXT NOT NULL,
                            action TEXT NOT NULL,
                            status TEXT NOT NULL,
                            message TEXT NOT NULL,
                            created_at TEXT NOT NULL
                        )
                        """
                    )
                    conn.commit()
                finally:
                    conn.close()

                init_analytics_db()
                record_analysis_run(
                    request_id="req_legacy",
                    user_id="user_legacy",
                    username="legacy-user",
                    session_id=None,
                    file_type="json",
                    status="success",
                    duration_ms=42,
                    holdings_count=2,
                    total_market_value=1000.0,
                )

                overview = get_admin_overview()
                self.assertEqual(len(overview["recent_analyses"]), 1)
                self.assertEqual(overview["recent_analyses"][0]["request_id"], "req_legacy")
                self.assertEqual(overview["recent_analyses"][0]["holdings_count"], 2)
                self.assertEqual(overview["recent_logs"][0]["metadata"]["request_id"], "req_legacy")
        finally:
            analytics_module.DB_PATH = original_db_path


class TestSupabaseAuthorization(unittest.TestCase):
    def test_supabase_url_allows_supabase_cloud_https_and_loopback_only(self):
        self.assertTrue(_is_allowed_supabase_url("https://project.supabase.co"))
        self.assertTrue(_is_allowed_supabase_url("https://project.supabase.co/"))
        self.assertTrue(_is_allowed_supabase_url("http://127.0.0.1:54321"))
        self.assertTrue(_is_allowed_supabase_url("http://localhost:54321"))
        self.assertTrue(_is_allowed_supabase_url("https://localhost:54321"))
        self.assertFalse(_is_allowed_supabase_url("https://example.com"))
        self.assertFalse(_is_allowed_supabase_url("https://project.supabase.co.evil.com"))
        self.assertFalse(_is_allowed_supabase_url("http://example.com"))
        self.assertFalse(_is_allowed_supabase_url("file:///tmp/auth"))

        with patch.dict(os.environ, {"SUPABASE_URL": "https://example.com"}, clear=False):
            self.assertEqual(_get_supabase_url(), "")

    def test_admin_check_ignores_user_mutable_metadata_roles(self):
        user = {
            "id": "user_123",
            "email": "user@example.com",
            "app_metadata": {},
            "user_metadata": {
                "role": "admin",
                "roles": ["admin"],
                "user_role": "admin",
                "user_roles": "admin",
            },
        }

        with patch.dict(
            os.environ,
            {
                "SUPABASE_ADMIN_ROLE": "admin",
                "SUPABASE_ADMIN_USER_IDS": "",
                "SUPABASE_ADMIN_EMAILS": "",
            },
            clear=False,
        ):
            self.assertFalse(_is_admin_user(user))

    def test_admin_check_accepts_trusted_app_metadata_role(self):
        user = {
            "id": "user_123",
            "email": "user@example.com",
            "app_metadata": {"role": "admin"},
            "user_metadata": {},
        }

        with patch.dict(
            os.environ,
            {
                "SUPABASE_ADMIN_ROLE": "admin",
                "SUPABASE_ADMIN_USER_IDS": "",
                "SUPABASE_ADMIN_EMAILS": "",
            },
            clear=False,
        ):
            self.assertTrue(_is_admin_user(user))

    def test_admin_check_requires_explicit_true_is_admin_claim(self):
        base_user = {
            "id": "user_123",
            "email": "user@example.com",
            "user_metadata": {},
        }

        false_values = [False, "false", "0", "no", "off", "", 0, 2, 0.0, 2.0]
        true_values = [True, "true", "1", "yes", 1, 1.0]

        with patch.dict(
            os.environ,
            {
                "SUPABASE_ADMIN_ROLE": "admin",
                "SUPABASE_ADMIN_USER_IDS": "",
                "SUPABASE_ADMIN_EMAILS": "",
            },
            clear=False,
        ):
            for value in false_values:
                with self.subTest(value=value):
                    user = {**base_user, "app_metadata": {"is_admin": value}}
                    self.assertFalse(_is_admin_user(user))

            for value in true_values:
                with self.subTest(value=value):
                    user = {**base_user, "app_metadata": {"is_admin": value}}
                    self.assertTrue(_is_admin_user(user))

    def test_username_sanitizer_redacts_common_pii(self):
        username = _sanitize_username("  Alice ABCDE1234F\nalice@example.com\t+91 98765 43210  ")

        self.assertEqual(
            username,
            "Alice [redacted-pan] [redacted-email] [redacted-phone]",
        )

    def test_username_redaction_migration_updates_existing_profiles(self):
        migration_sql = Path(
            "supabase/migrations/20260605000000_harden_username_pii_redaction.sql"
        ).read_text(encoding="utf-8")
        normalized_sql = re.sub(r"\s+", " ", migration_sql.lower())

        self.assertIn("create or replace function public.echo_clean_username", normalized_sql)
        self.assertIn("[redacted-pan]", normalized_sql)
        self.assertIn("[redacted-email]", normalized_sql)
        self.assertIn("[redacted-phone]", normalized_sql)
        self.assertIn("update public.profiles set username = public.echo_clean_username(username, id)", normalized_sql)


class TestBillingSecurity(unittest.TestCase):
    def test_admin_users_get_unlimited_access_without_rpc(self):
        from app.Code.billing import (
            admin_unlimited_access_status,
            get_access_status,
            reserve_analysis_credit,
        )

        async def never_called(*args, **kwargs):
            raise AssertionError("Supabase RPC should not be called for admin users")

        with patch("app.Code.billing._supabase_rpc", new=never_called):
            access = asyncio.run(get_access_status("user_admin", is_admin=True))
            reservation = asyncio.run(
                reserve_analysis_credit("user_admin", is_admin=True)
            )

        expected = admin_unlimited_access_status()
        self.assertTrue(access.can_analyze)
        self.assertTrue(access.has_unlimited_reports)
        self.assertEqual(access.subscription_status, "active")
        self.assertTrue(reservation.allowed)
        self.assertFalse(reservation.credit_consumed)
        self.assertEqual(reservation.access.to_dict(), expected.to_dict())

    def test_billing_access_endpoint_returns_unlimited_for_admin_without_rpc(self):
        client = TestClient(app)

        async def never_called(*args, **kwargs):
            raise AssertionError("Supabase RPC should not be called for admin users")

        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.billing._supabase_rpc",
            new=never_called,
        ):
            response = client.get("/api/billing/access")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["has_unlimited_reports"])
        self.assertTrue(body["can_analyze"])
        self.assertEqual(body["subscription_status"], "active")

    def test_admin_parse_pdf_skips_credit_consumption_without_rpc(self):
        client = TestClient(app)
        parser_payload = {"folios": []}

        async def never_called(*args, **kwargs):
            raise AssertionError("Supabase RPC should not be called for admin users")

        refund_mock = AsyncMock()
        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.billing._supabase_rpc",
            new=never_called,
        ), patch(
            "app.Code.main.refund_analysis_credit",
            new=refund_mock,
        ), patch(
            "app.Code.main._parse_pdf_upload",
            new=AsyncMock(return_value={"success": True, "data": parser_payload}),
        ):
            response = client.post(
                "/api/parse_pdf",
                files={"file": ("statement.pdf", b"%PDF-1.7\n", "application/pdf")},
                data={"password": "", "output_format": "json"},
            )

        self.assertEqual(response.status_code, 200)
        refund_mock.assert_not_awaited()

    def test_billing_access_and_reservation_parsing_fail_closed(self):
        access = _parse_access_status({})
        self.assertFalse(access.can_analyze)
        self.assertEqual(access.cas_report_limit, 0)
        self.assertEqual(access.remaining_free_reports, 0)

        reservation = _parse_credit_reservation([])
        self.assertFalse(reservation.allowed)
        self.assertFalse(reservation.credit_consumed)

        reservation = _parse_credit_reservation(
            {
                "allowed": False,
                "can_analyze": True,
                "has_unlimited_reports": False,
                "cas_report_limit": 1,
                "cas_reports_used": 0,
                "remaining_free_reports": 1,
                "subscription_status": "free",
            }
        )
        self.assertFalse(reservation.allowed)

    def test_razorpay_checkout_signature_validates_ids_and_signature(self):
        with patch.dict(os.environ, {"RAZORPAY_KEY_SECRET": "test_secret"}, clear=False):
            signature = hmac.new(
                b"test_secret",
                b"pay_test123|sub_test123",
                hashlib.sha256,
            ).hexdigest()
            verify_checkout_signature(
                payment_id="pay_test123",
                subscription_id="sub_test123",
                signature=signature,
            )

            with self.assertRaises(HTTPException) as invalid_id:
                verify_checkout_signature(
                    payment_id="pay_test123",
                    subscription_id="sub_test123/../../other",
                    signature=signature,
                )
            self.assertEqual(invalid_id.exception.status_code, 400)

            with self.assertRaises(HTTPException) as invalid_signature:
                verify_checkout_signature(
                    payment_id="pay_test123",
                    subscription_id="sub_test123",
                    signature="not-a-hex-signature",
                )
            self.assertEqual(invalid_signature.exception.status_code, 400)

    def test_apply_subscription_status_rejects_invalid_state_before_rpc(self):
        async def never_called(*args, **kwargs):
            raise AssertionError("Supabase RPC should not be called")

        with patch("app.Code.billing._supabase_rpc", new=never_called):
            with self.assertRaises(HTTPException) as bad_id:
                asyncio.run(
                    apply_subscription_status(
                        user_id="user_test",
                        subscription_id="sub_test123/../../other",
                        customer_id=None,
                        status="active",
                        current_period_end=None,
                    )
                )
            self.assertEqual(bad_id.exception.status_code, 400)

            with self.assertRaises(HTTPException) as bad_status:
                asyncio.run(
                    apply_subscription_status(
                        user_id="user_test",
                        subscription_id="sub_test123",
                        customer_id=None,
                        status="trialing",
                        current_period_end=None,
                    )
                )
            self.assertEqual(bad_status.exception.status_code, 400)

    def test_razorpay_webhook_signature_uses_exact_raw_body(self):
        raw_body = b'{"event":"subscription.activated","payload":{}}'
        with patch.dict(os.environ, {"RAZORPAY_WEBHOOK_SECRET": "webhook_secret"}, clear=False):
            signature = hmac.new(b"webhook_secret", raw_body, hashlib.sha256).hexdigest()
            verify_webhook_signature(raw_body, signature)

            with self.assertRaises(HTTPException) as mismatched_body:
                verify_webhook_signature(raw_body + b"\n", signature)
            self.assertEqual(mismatched_body.exception.status_code, 400)

    def test_razorpay_webhook_rejects_oversized_body_before_signature_work(self):
        client = TestClient(app)
        with patch("app.Code.main.MAX_WEBHOOK_BODY_BYTES", 10), patch(
            "app.Code.main.verify_webhook_signature",
            side_effect=AssertionError("signature should not be checked for oversized bodies"),
        ):
            response = client.post(
                "/api/billing/razorpay-webhook",
                content=b"x" * 11,
                headers={"x-razorpay-signature": "0" * 64},
            )

        self.assertEqual(response.status_code, 413)
        self.assertIn("too large", response.json().get("detail", "").lower())

    def test_razorpay_webhook_rejects_invalid_event_id_before_rpc(self):
        client = TestClient(app)
        has_event_mock = AsyncMock()
        apply_mock = AsyncMock()
        payload = {"event": "subscription.activated", "payload": {}}

        with patch("app.Code.main.verify_webhook_signature"), patch(
            "app.Code.main.has_webhook_event",
            new=has_event_mock,
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=apply_mock,
        ):
            response = client.post(
                "/api/billing/razorpay-webhook",
                content=json.dumps(payload).encode("utf-8"),
                headers={
                    "x-razorpay-signature": "0" * 64,
                    "x-razorpay-event-id": "evt_" + ("x" * 200),
                },
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("event id", response.json().get("detail", "").lower())
        has_event_mock.assert_not_awaited()
        apply_mock.assert_not_awaited()

    def test_razorpay_webhook_duplicate_event_skips_subscription_apply(self):
        client = TestClient(app)
        apply_mock = AsyncMock()
        claim_mock = AsyncMock()
        payload = {
            "event": "subscription.activated",
            "payload": {
                "subscription": {
                    "entity": {
                        "id": "sub_test123",
                        "plan_id": "plan_good123",
                        "status": "active",
                        "customer_id": "cust_test123",
                        "current_end": 1893456000,
                        "notes": {"user_id": "user_test"},
                    }
                }
            },
        }

        with patch.dict(os.environ, {"RAZORPAY_PLAN_ID": "plan_good123"}, clear=False), patch(
            "app.Code.main.verify_webhook_signature",
        ), patch(
            "app.Code.main.has_webhook_event",
            new=AsyncMock(return_value=True),
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=apply_mock,
        ), patch(
            "app.Code.main.claim_webhook_event",
            new=claim_mock,
        ):
            response = client.post(
                "/api/billing/razorpay-webhook",
                content=json.dumps(payload).encode("utf-8"),
                headers={
                    "x-razorpay-signature": "0" * 64,
                    "x-razorpay-event-id": "evt_test123",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("duplicate"))
        apply_mock.assert_not_awaited()
        claim_mock.assert_not_awaited()

    def test_razorpay_webhook_does_not_claim_event_when_subscription_apply_fails(self):
        client = TestClient(app)
        claim_mock = AsyncMock()
        payload = {
            "event": "subscription.activated",
            "payload": {
                "subscription": {
                    "entity": {
                        "id": "sub_test123",
                        "plan_id": "plan_good123",
                        "status": "active",
                        "customer_id": "cust_test123",
                        "current_end": 1893456000,
                        "notes": {"user_id": "user_test"},
                    }
                }
            },
        }

        with patch.dict(os.environ, {"RAZORPAY_PLAN_ID": "plan_good123"}, clear=False), patch(
            "app.Code.main.verify_webhook_signature",
        ), patch(
            "app.Code.main.has_webhook_event",
            new=AsyncMock(return_value=False),
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=AsyncMock(side_effect=HTTPException(status_code=503, detail="Supabase unavailable.")),
        ), patch(
            "app.Code.main.claim_webhook_event",
            new=claim_mock,
        ):
            response = client.post(
                "/api/billing/razorpay-webhook",
                content=json.dumps(payload).encode("utf-8"),
                headers={
                    "x-razorpay-signature": "0" * 64,
                    "x-razorpay-event-id": "evt_test123",
                },
            )

        self.assertEqual(response.status_code, 503)
        claim_mock.assert_not_awaited()

    def test_subscription_webhook_event_requires_configured_plan(self):
        payload = {
            "payload": {
                "subscription": {
                    "entity": {
                        "id": "sub_test123",
                        "plan_id": "plan_good123",
                        "status": "active",
                        "customer_id": "cust_test123",
                        "current_end": 1893456000,
                        "notes": {"user_id": "user_test"},
                    }
                }
            }
        }

        with patch.dict(os.environ, {"RAZORPAY_PLAN_ID": "plan_good123"}, clear=False):
            event = extract_subscription_event(payload)
        self.assertIsNotNone(event)
        self.assertEqual(event["subscription_id"], "sub_test123")
        self.assertEqual(event["status"], "active")
        self.assertEqual(event["user_id"], "user_test")

        with patch.dict(os.environ, {"RAZORPAY_PLAN_ID": "plan_other123"}, clear=False):
            self.assertIsNone(extract_subscription_event(payload))

        bad_status_payload = json.loads(json.dumps(payload))
        bad_status_payload["payload"]["subscription"]["entity"]["status"] = "trialing"
        with patch.dict(os.environ, {"RAZORPAY_PLAN_ID": "plan_good123"}, clear=False):
            self.assertIsNone(extract_subscription_event(bad_status_payload))

    def test_supabase_billing_migration_restricts_profile_writes_and_rpc_execute(self):
        hardening_sql = Path(
            "supabase/migrations/20260604000000_harden_profile_billing_permissions.sql"
        ).read_text(encoding="utf-8")
        normalized_sql = re.sub(r"\s+", " ", hardening_sql.lower())

        self.assertIn("revoke update on table public.profiles from anon, authenticated", normalized_sql)
        self.assertIn("grant update (username) on table public.profiles to authenticated", normalized_sql)

        privileged_functions = [
            "public.echo_get_access_status(uuid)",
            "public.echo_consume_report_credit(uuid)",
            "public.echo_refund_report_credit(uuid)",
            "public.echo_apply_razorpay_subscription_event(uuid, text, text, text, timestamptz)",
            "public.echo_claim_razorpay_webhook_event(text, text)",
        ]
        for function_signature in privileged_functions:
            with self.subTest(function_signature=function_signature):
                self.assertIn(
                    f"revoke execute on function {function_signature} from public, anon, authenticated",
                    normalized_sql,
                )
                self.assertIn(
                    f"grant execute on function {function_signature} to service_role",
                    normalized_sql,
                )

    def test_razorpay_webhook_idempotency_migration_creates_backend_only_check(self):
        migration_sql = Path(
            "supabase/migrations/20260604001000_harden_razorpay_webhook_idempotency.sql"
        ).read_text(encoding="utf-8")
        normalized_sql = re.sub(r"\s+", " ", migration_sql.lower())

        self.assertIn("create or replace function public.echo_has_razorpay_webhook_event(event_id text)", normalized_sql)
        self.assertIn(
            "revoke execute on function public.echo_has_razorpay_webhook_event(text) from public, anon, authenticated",
            normalized_sql,
        )
        self.assertIn(
            "grant execute on function public.echo_has_razorpay_webhook_event(text) to service_role",
            normalized_sql,
        )
        self.assertIn("public.profiles.subscription_status in ('authenticated', 'active')", normalized_sql)
        self.assertIn("in ('created', 'pending')", normalized_sql)
        self.assertIn("razorpay_subscription_current_end = case", normalized_sql)
        self.assertIn("else coalesce(new_current_period_end, public.profiles.razorpay_subscription_current_end)", normalized_sql)

    def test_create_subscription_refuses_when_access_is_already_unlimited(self):
        client = TestClient(app)
        active_access = AccessStatus(
            can_analyze=True,
            has_unlimited_reports=True,
            cas_report_limit=1,
            cas_reports_used=1,
            remaining_free_reports=0,
            subscription_status="active",
            razorpay_subscription_id="sub_test123",
            current_period_end=None,
        )

        with patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.get_access_status",
            new=AsyncMock(return_value=active_access),
        ), patch(
            "app.Code.main.create_razorpay_subscription",
            new=AsyncMock(side_effect=AssertionError("Razorpay should not be called")),
        ):
            response = client.post("/api/billing/create-subscription")

        self.assertEqual(response.status_code, 409)
        self.assertIn("already active", response.json().get("detail", "").lower())

    def test_create_subscription_replaces_stale_created_checkout_subscription(self):
        client = TestClient(app)
        stale_access = _test_access_status(
            can_analyze=False,
            cas_reports_used=1,
            remaining_free_reports=0,
            subscription_status="created",
            razorpay_subscription_id="sub_old123",
        )
        refreshed_access = _test_access_status(
            can_analyze=False,
            cas_reports_used=1,
            remaining_free_reports=0,
            subscription_status="created",
            razorpay_subscription_id="sub_new123",
        )
        fetch_subscription = AsyncMock(
            return_value={
                "id": "sub_old123",
                "status": "expired",
                "plan_id": "plan_good123",
                "customer_id": "cust_test123",
                "current_end": None,
                "notes": {"user_id": "user_test"},
            }
        )
        apply_status = AsyncMock(
            return_value=_test_access_status(
                can_analyze=False,
                cas_reports_used=1,
                remaining_free_reports=0,
                subscription_status="expired",
                razorpay_subscription_id="sub_old123",
            )
        )
        create_subscription = AsyncMock(
            return_value={
                "id": "sub_new123",
                "status": "created",
                "plan_id": "plan_good123",
            }
        )

        with patch.dict(
            os.environ,
            {"RAZORPAY_KEY_ID": "rzp_test_key", "RAZORPAY_PLAN_ID": "plan_good123"},
            clear=False,
        ), patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.get_access_status",
            new=AsyncMock(side_effect=[stale_access, refreshed_access]),
        ), patch(
            "app.Code.main.fetch_razorpay_subscription",
            new=fetch_subscription,
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=apply_status,
        ), patch(
            "app.Code.main.create_razorpay_subscription",
            new=create_subscription,
        ):
            response = client.post("/api/billing/create-subscription")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["subscription_id"], "sub_new123")
        fetch_subscription.assert_awaited_once_with("sub_old123")
        apply_status.assert_awaited_once()
        create_subscription.assert_awaited_once()

    def test_create_subscription_does_not_send_username_to_razorpay_notes(self):
        captured_request = {}

        class CapturingRazorpayClient:
            def __init__(self, *args, **kwargs):
                pass

            async def __aenter__(self):
                return self

            async def __aexit__(self, exc_type, exc, tb):
                return False

            async def post(self, url, **kwargs):
                captured_request["url"] = url
                captured_request["json"] = kwargs.get("json")
                return _FakeResponse(
                    200,
                    {
                        "id": "sub_test123",
                        "status": "created",
                        "plan_id": "plan_good123",
                        "customer_id": "cust_test123",
                        "current_end": 1893456000,
                    },
                )

        with patch.dict(
            os.environ,
            {
                "RAZORPAY_KEY_ID": "rzp_test_key",
                "RAZORPAY_KEY_SECRET": "test_secret",
                "RAZORPAY_PLAN_ID": "plan_good123",
            },
            clear=False,
        ), patch(
            "app.Code.billing.httpx.AsyncClient",
            CapturingRazorpayClient,
        ), patch(
            "app.Code.billing.apply_subscription_status",
            new=AsyncMock(return_value=_test_access_status()),
        ):
            subscription = asyncio.run(create_razorpay_subscription(_FakeSupabaseUser()))

        self.assertEqual(subscription["id"], "sub_test123")
        notes = captured_request["json"]["notes"]
        self.assertEqual(notes["user_id"], "user_test")
        self.assertEqual(notes["source"], "echo-analyze")
        self.assertNotIn("username", notes)

    def test_verify_subscription_rejects_wrong_plan_before_applying_access(self):
        client = TestClient(app)
        access = AccessStatus(
            can_analyze=False,
            has_unlimited_reports=False,
            cas_report_limit=1,
            cas_reports_used=1,
            remaining_free_reports=0,
            subscription_status="created",
            razorpay_subscription_id="sub_test123",
            current_period_end=None,
        )
        signature = hmac.new(
            b"test_secret",
            b"pay_test123|sub_test123",
            hashlib.sha256,
        ).hexdigest()

        apply_mock = AsyncMock()
        with patch.dict(
            os.environ,
            {"RAZORPAY_KEY_SECRET": "test_secret", "RAZORPAY_PLAN_ID": "plan_good123"},
            clear=False,
        ), patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.get_access_status",
            new=AsyncMock(return_value=access),
        ), patch(
            "app.Code.main.fetch_razorpay_subscription",
            new=AsyncMock(return_value={"id": "sub_test123", "status": "active", "plan_id": "plan_bad123"}),
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=apply_mock,
        ):
            response = client.post(
                "/api/billing/verify-subscription-payment",
                json={
                    "razorpay_payment_id": "pay_test123",
                    "razorpay_subscription_id": "sub_test123",
                    "razorpay_signature": signature,
                },
            )

        self.assertEqual(response.status_code, 400)
        apply_mock.assert_not_called()

    def test_verify_subscription_payment_rejects_oversized_fields(self):
        client = TestClient(app)
        response = client.post(
            "/api/billing/verify-subscription-payment",
            json={
                "razorpay_payment_id": "pay_" + ("x" * 80),
                "razorpay_subscription_id": "sub_test123",
                "razorpay_signature": "0" * 64,
            },
        )

        self.assertEqual(response.status_code, 422)

    def test_verify_subscription_rejects_missing_status_before_applying_access(self):
        client = TestClient(app)
        access = AccessStatus(
            can_analyze=False,
            has_unlimited_reports=False,
            cas_report_limit=1,
            cas_reports_used=1,
            remaining_free_reports=0,
            subscription_status="created",
            razorpay_subscription_id="sub_test123",
            current_period_end=None,
        )
        signature = hmac.new(
            b"test_secret",
            b"pay_test123|sub_test123",
            hashlib.sha256,
        ).hexdigest()

        apply_mock = AsyncMock()
        with patch.dict(
            os.environ,
            {"RAZORPAY_KEY_SECRET": "test_secret", "RAZORPAY_PLAN_ID": "plan_good123"},
            clear=False,
        ), patch("app.Code.main.require_supabase_user", new=_fake_require_supabase_user), patch(
            "app.Code.main.get_access_status",
            new=AsyncMock(return_value=access),
        ), patch(
            "app.Code.main.fetch_razorpay_subscription",
            new=AsyncMock(return_value={"id": "sub_test123", "plan_id": "plan_good123"}),
        ), patch(
            "app.Code.main.apply_subscription_status",
            new=apply_mock,
        ):
            response = client.post(
                "/api/billing/verify-subscription-payment",
                json={
                    "razorpay_payment_id": "pay_test123",
                    "razorpay_subscription_id": "sub_test123",
                    "razorpay_signature": signature,
                },
            )

        self.assertEqual(response.status_code, 502)
        apply_mock.assert_not_called()


class TestParserAndHoldingsResilience(unittest.IsolatedAsyncioTestCase):
    def test_pdf_parse_timeout_default_and_cap(self):
        with patch.dict(os.environ, {"PDF_PARSE_TIMEOUT_SECONDS": ""}, clear=False):
            self.assertEqual(_get_pdf_parse_timeout_seconds(), 120.0)

        with patch.dict(os.environ, {"PDF_PARSE_TIMEOUT_SECONDS": "999"}, clear=False):
            self.assertEqual(_get_pdf_parse_timeout_seconds(), 240.0)

        with patch.dict(os.environ, {"PDF_PARSE_TIMEOUT_SECONDS": "invalid"}, clear=False):
            self.assertEqual(_get_pdf_parse_timeout_seconds(), 120.0)

    def test_pdf_parse_auto_uses_thread_on_vercel(self):
        with patch.dict(os.environ, {"PDF_PARSE_EXECUTOR": "auto", "VERCEL": "1"}, clear=False):
            self.assertEqual(_get_pdf_parse_executor(), "thread")

        with patch.dict(os.environ, {"PDF_PARSE_EXECUTOR": "auto", "VERCEL_ENV": "production"}, clear=False):
            os.environ.pop("VERCEL", None)
            self.assertEqual(_get_pdf_parse_executor(), "thread")

    def test_pdf_parse_auto_uses_process_locally(self):
        with patch.dict(os.environ, {"PDF_PARSE_EXECUTOR": "auto"}, clear=False):
            os.environ.pop("VERCEL", None)
            os.environ.pop("VERCEL_ENV", None)
            self.assertEqual(_get_pdf_parse_executor(), "auto")

    def test_pdf_parse_explicit_process_overrides_vercel_default(self):
        with patch.dict(os.environ, {"PDF_PARSE_EXECUTOR": "process", "VERCEL": "1"}, clear=False):
            self.assertEqual(_get_pdf_parse_executor(), "process")

    def test_pdf_parse_subprocess_timeout_terminates_worker(self):
        started_at = time.perf_counter()

        result = _parse_pdf_upload_in_subprocess(
            b"%PDF-1.4\n",
            "",
            0.2,
            worker_target=_sleeping_pdf_parse_worker,
        )

        self.assertFalse(result["success"])
        self.assertIn("timed out", result["error"].lower())
        self.assertLess(time.perf_counter() - started_at, 4.0)

    def test_pdf_parse_subprocess_reads_worker_pipe_result(self):
        result = _parse_pdf_upload_in_subprocess(
            b"%PDF-1.4\n",
            "",
            5.0,
            worker_target=_successful_pdf_parse_worker,
        )

        self.assertTrue(result["success"])
        self.assertEqual(result["data"], {"folios": []})

    def test_pdf_parse_subprocess_setup_failure_is_controlled_error(self):
        class BrokenContext:
            def Pipe(self, *args, **kwargs):
                raise OSError("process pipes unavailable")

        with patch("app.Code.main.multiprocessing.get_context", return_value=BrokenContext()):
            result = _parse_pdf_upload_in_subprocess(b"%PDF-1.4\n", "", 0.2)

        self.assertFalse(result["success"])
        self.assertIn("could not start", result["error"].lower())

    async def test_pdf_parse_auto_falls_back_to_thread_when_subprocess_cannot_start(self):
        class BrokenContext:
            def Pipe(self, *args, **kwargs):
                raise OSError("process pipes unavailable")

        with patch.dict(
            os.environ,
            {"PDF_PARSE_EXECUTOR": "auto", "PDF_PARSE_TIMEOUT_SECONDS": "1"},
            clear=False,
        ), patch(
            "app.Code.main.multiprocessing.get_context",
            return_value=BrokenContext(),
        ), patch(
            "app.Code.main.parse_with_casparser",
            return_value={"success": True, "data": {"folios": []}},
        ):
            os.environ.pop("VERCEL", None)
            os.environ.pop("VERCEL_ENV", None)
            result = await _parse_pdf_upload(b"%PDF-1.4\n", password="")

        self.assertTrue(result["success"])
        self.assertEqual(result["data"], {"folios": []})

    async def test_pdf_parse_thread_executor_skips_subprocess(self):
        with patch.dict(
            os.environ,
            {"PDF_PARSE_EXECUTOR": "thread", "PDF_PARSE_TIMEOUT_SECONDS": "1"},
            clear=False,
        ), patch(
            "app.Code.main.multiprocessing.get_context",
            side_effect=AssertionError("subprocess should not start"),
        ), patch(
            "app.Code.main.parse_with_casparser",
            return_value={"success": True, "data": {"folios": []}},
        ):
            result = await _parse_pdf_upload(b"%PDF-1.4\n", password="")

        self.assertTrue(result["success"])
        self.assertEqual(result["data"], {"folios": []})

    def test_pdf_parse_subprocess_reads_large_result_before_joining_worker(self):
        class FakeConnection:
            def __init__(self):
                self.drained = False

            def poll(self, timeout=None):
                return True

            def recv(self):
                self.drained = True
                return {"success": True, "data": {"folios": []}}

            def close(self):
                return None

        class FakeProcess:
            def __init__(self, target, args, daemon):
                self.connection = args[2]
                self.started = False
                self.terminated = False

            def start(self):
                self.started = True

            def join(self, timeout=None):
                return None

            def is_alive(self):
                return not self.connection.drained and not self.terminated

            def terminate(self):
                self.terminated = True

            def kill(self):
                self.terminated = True

        class FakeContext:
            def __init__(self):
                self.parent_connection = FakeConnection()
                self.child_connection = self.parent_connection

            def Pipe(self, *args, **kwargs):
                return self.parent_connection, self.child_connection

            def Process(self, target, args, daemon):
                return FakeProcess(target, args, daemon)

        with patch("app.Code.main.multiprocessing.get_context", return_value=FakeContext()):
            result = _parse_pdf_upload_in_subprocess(b"%PDF-1.4\n", "", 0.2)

        self.assertTrue(result["success"])
        self.assertEqual(result["data"], {"folios": []})

    def test_pdfminer_cmap_loader_rejects_path_like_names(self):
        from pdfminer.cmapdb import CMapDB

        harden_pdfminer_cmap_loading()

        with patch.object(CMapDB, ORIGINAL_CMAP_LOADER_ATTR, side_effect=AssertionError("unsafe loader called")) as original:
            for name in [
                "../evil",
                "..\\evil",
                "/tmp/evil",
                "C:\\tmp\\evil",
                "\\\\server\\share\\evil",
                "Adobe/../../evil",
                "Identity-H\0",
                "",
            ]:
                with self.subTest(name=name):
                    with self.assertRaises(CMapDB.CMapNotFound):
                        CMapDB._load_data(name)

        original.assert_not_called()

    def test_pdfminer_cmap_loader_still_delegates_safe_names(self):
        from pdfminer.cmapdb import CMapDB

        harden_pdfminer_cmap_loading()

        with patch.object(CMapDB, ORIGINAL_CMAP_LOADER_ATTR, return_value={"ok": True}) as original:
            self.assertEqual(CMapDB._load_data("Identity-H"), {"ok": True})

        original.assert_called_once_with("Identity-H")

    def test_parse_with_casparser_reapplies_pdfminer_hardening(self):
        with patch("app.Code.cas_parser.harden_pdfminer_cmap_loading") as harden, patch(
            "app.Code.cas_parser.read_cas_pdf", return_value={"folios": []}
        ):
            result = parse_with_casparser("dummy.pdf")

        self.assertTrue(result["success"])
        harden.assert_called_once()

    def test_parse_with_casparser_path_input_skips_tempfile_creation(self):
        with patch("app.Code.cas_parser.read_cas_pdf", return_value={"folios": []}), patch(
            "tempfile.NamedTemporaryFile"
        ) as named_tmp:
            result = parse_with_casparser("dummy.pdf")

        self.assertTrue(result["success"])
        named_tmp.assert_not_called()

    def test_pdf_parse_direct_does_not_swallow_base_exceptions(self):
        class FatalParserExit(BaseException):
            pass

        with patch("app.Code.main.parse_with_casparser", side_effect=FatalParserExit):
            with self.assertRaises(FatalParserExit):
                _parse_pdf_upload_direct(b"%PDF-1.4\n", "")

    def test_recursive_to_dict_serializes_json_unsafe_parser_values(self):
        converted = recursive_to_dict(
            {
                "amount": Decimal("100.50"),
                "date": date(2024, 1, 2),
                "type": TransactionType.PURCHASE,
            }
        )

        json.dumps(converted)
        self.assertEqual(converted["amount"], "100.50")
        self.assertEqual(converted["date"], "2024-01-02")
        self.assertEqual(converted["type"], TransactionType.PURCHASE.value)

    async def test_groww_index_failure_does_not_disable_future_retries(self):
        old_loaded = holdings_module._groww_index_loaded
        old_by_code = holdings_module._groww_index_by_code
        old_entries = holdings_module._groww_index_entries
        try:
            holdings_module._groww_index_loaded = False
            holdings_module._groww_index_by_code = {}
            holdings_module._groww_index_entries = []

            class FailingClient:
                async def get(self, url):
                    return _FakeResponse(500, {})

            by_code, entries = await holdings_module._get_groww_index(FailingClient())
            self.assertEqual(by_code, {})
            self.assertEqual(entries, [])
            self.assertFalse(holdings_module._groww_index_loaded)
        finally:
            holdings_module._groww_index_loaded = old_loaded
            holdings_module._groww_index_by_code = old_by_code
            holdings_module._groww_index_entries = old_entries

    async def test_amfi_transient_failure_does_not_persist_failed_urls(self):
        old_cache = holdings_module._amfi_cache
        old_failed = holdings_module._failed_urls
        try:
            holdings_module._amfi_cache = {}
            holdings_module._failed_urls = set()

            with patch("app.Code.holdings.httpx.AsyncClient", _FailingAsyncClient), patch(
                "app.Code.holdings.save_amfi_cache_async", new=AsyncMock()
            ):
                result = await holdings_module._fetch_amfi_monthly_file()

            self.assertEqual(result, {})
            self.assertEqual(holdings_module._failed_urls, set())
        finally:
            holdings_module._amfi_cache = old_cache
            holdings_module._failed_urls = old_failed

    async def test_pdf_upload_parser_times_out_without_blocking_event_loop(self):
        def slow_parse(*_args, **_kwargs):
            time.sleep(0.1)
            return {"success": True, "data": {"folios": []}}

        with patch("app.Code.main._get_pdf_parse_timeout_seconds", return_value=0.01), patch(
            "app.Code.main.parse_with_casparser", side_effect=slow_parse
        ):
            result = await _parse_pdf_upload(b"%PDF-1.7\n", password="")

        self.assertFalse(result["success"])
        self.assertIn("timed out", result["error"].lower())

    def test_convert_to_excel_escapes_formula_like_cells(self):
        payload = {
            "folios": [
                {
                    "amc": "=AMC",
                    "folio": "+12345",
                    "schemes": [
                        {
                            "scheme": "@Danger Fund",
                            "advisor": "-Advisor",
                            "transactions": [
                                {
                                    "date": "2024-01-01",
                                    "description": "=cmd",
                                    "amount": 1000,
                                    "units": 10,
                                    "nav": 100,
                                    "balance": 10,
                                    "type": "@TYPE",
                                },
                                {
                                    "date": "2024-01-02",
                                    "description": " \t=cmd",
                                    "amount": 1000,
                                    "units": 10,
                                    "nav": 100,
                                    "balance": 10,
                                    "type": "\t@TYPE",
                                },
                                {
                                    "date": "2024-01-03",
                                    "description": "\nplain",
                                    "amount": 1000,
                                    "units": 10,
                                    "nav": 100,
                                    "balance": 10,
                                    "type": "\nTYPE",
                                }
                            ],
                        }
                    ],
                }
            ]
        }

        workbook_bytes = convert_to_excel(payload)
        workbook = load_workbook(workbook_bytes)
        sheet = workbook.active

        self.assertEqual(sheet["A2"].value, "'=AMC")
        self.assertEqual(sheet["B2"].value, "'+12345")
        self.assertEqual(sheet["C2"].value, "'@Danger Fund")
        self.assertEqual(sheet["D2"].value, "'-Advisor")
        self.assertEqual(sheet["F2"].value, "'=cmd")
        self.assertEqual(sheet["K2"].value, "'@TYPE")
        self.assertEqual(sheet["F3"].value, "' \t=cmd")
        self.assertEqual(sheet["K3"].value, "'\t@TYPE")
        self.assertEqual(sheet["F4"].value, "'\nplain")
        self.assertEqual(sheet["K4"].value, "'\nTYPE")

    def test_convert_to_excel_skips_malformed_rows_without_crashing(self):
        workbook_bytes = convert_to_excel(
            {
                "folios": [
                    "bad folio",
                    {
                        "schemes": [
                            "bad scheme",
                            {"transactions": ["bad transaction"]},
                        ],
                    },
                ]
            }
        )
        workbook = load_workbook(workbook_bytes)
        sheet = workbook.active

        self.assertEqual(sheet["A2"].value, "No transactions found")

    def test_vendored_casparser_csv_exports_escape_formula_like_cells(self):
        cas_data = CASData(
            statement_period=StatementPeriod(from_="01-Jan-2024", to="01-Jan-2025"),
            investor_info=CasInvestorInfo(
                name="Investor",
                email="investor@example.com",
                address="Address",
                mobile="9999999999",
            ),
            cas_type=CASFileType.DETAILED,
            file_type=FileType.CAMS,
            folios=[
                Folio(
                    folio="=12345",
                    amc="@AMC",
                    PAN="ABCDE1234F",
                    schemes=[
                        Scheme(
                            scheme="+Danger Fund",
                            advisor="-Advisor",
                            rta_code="RTA",
                            rta="CAMS",
                            type="EQUITY",
                            isin="INF000000000",
                            amfi="100001",
                            nominees=[],
                            open=Decimal("0"),
                            close=Decimal("10"),
                            close_calculated=Decimal("10"),
                            valuation=SchemeValuation(
                                date="2024-01-01",
                                nav=Decimal("100"),
                                value=Decimal("1000"),
                            ),
                            transactions=[
                                TransactionData(
                                    date="2024-01-01",
                                    description=" \t=cmd",
                                    amount=Decimal("1000"),
                                    units=Decimal("10"),
                                    nav=Decimal("100"),
                                    balance=Decimal("10"),
                                    type=TransactionType.PURCHASE,
                                )
                            ],
                        )
                    ],
                )
            ],
        )

        detailed_csv = cas2csv(cas_data)
        summary_csv = cas2csv_summary(cas_data)
        detailed_rows = list(csv.DictReader(io.StringIO(detailed_csv)))
        summary_rows = list(csv.DictReader(io.StringIO(summary_csv)))

        self.assertEqual(detailed_rows[0]["amc"], "'@AMC")
        self.assertEqual(detailed_rows[0]["folio"], "'=12345")
        self.assertEqual(detailed_rows[0]["scheme"], "'+Danger Fund")
        self.assertEqual(detailed_rows[0]["advisor"], "'-Advisor")
        self.assertEqual(detailed_rows[0]["description"], "' \t=cmd")
        self.assertEqual(summary_rows[0]["amc"], "'@AMC")
        self.assertEqual(summary_rows[0]["scheme"], "'+Danger Fund")
